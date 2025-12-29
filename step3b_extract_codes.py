import json
import os
from openpyxl import load_workbook
import win32com.client
import pythoncom

# =========================
# CONFIG
# =========================
MAP_FILE = "sheet_data/key_file_map.json"
OUT_DIR = "sheet_data/codes"
TARGET_SHEETS = ["MASTER SPEC "]

os.makedirs(OUT_DIR, exist_ok=True)

LABEL_MAP = {
    "ชนิดขวด": "ขวด_train",
    "ชนิดฝา": "ฝา_train",
    "ชนิดแคปซีล": "แคปซีล_train",
    "ฟอยด์": "ฟอยด์_train",
    "ฉลาก": "แบบฉลาก_train",
    "สติ๊กเกอร์": "สติ๊กเกอร์_train",   # ⭐ ตัวเดียว แต่หลายรหัส
}

# =========================
# UTIL
# =========================
def is_code(val):
    return isinstance(val, str) and val.strip().startswith("P")

def convert_xls_to_xlsx(xls_path):
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    xlsx_path = xls_path + "x"
    wb = excel.Workbooks.Open(xls_path)
    wb.SaveAs(xlsx_path, FileFormat=51)
    wb.Close(False)
    excel.Quit()

    print(f"🔄 converted: {os.path.basename(xls_path)} → {os.path.basename(xlsx_path)}")
    return xlsx_path

# =========================
# CORE
# =========================
def find_codes(ws):
    rows = list(ws.iter_rows(values_only=True))
    result = {}

    for i, row in enumerate(rows):
        if not row or not row[0]:
            continue

        left = str(row[0]).strip()

        # =========================
        # กล่อง / ถาด (รวม class เดียว)
        # =========================
        if left.startswith("4.1") or left.startswith("4.2"):
            for j in range(i, i + 12):
                if j >= len(rows):
                    break
                for cell in rows[j]:
                    if is_code(cell):
                        result.setdefault("แบบถาด-กล่อง_train", cell.strip())
                        break
                if "แบบถาด-กล่อง_train" in result:
                    break

        # =========================
        # class อื่น ๆ
        # =========================
        for label, cls in LABEL_MAP.items():
            if cls != "สติ๊กเกอร์_train" and cls in result:
                continue

            if label in left:
                for cell in row:
                    if is_code(cell):
                        if cls == "สติ๊กเกอร์_train":
                            result.setdefault(cls, []).append(cell.strip())
                        else:
                            result[cls] = cell.strip()
                        break

                # เผื่อ merged cell
                if cls not in result or cls == "สติ๊กเกอร์_train":
                    for k in range(1, 3):
                        if i + k < len(rows):
                            for cell in rows[i + k]:
                                if is_code(cell):
                                    if cls == "สติ๊กเกอร์_train":
                                        result.setdefault(cls, []).append(cell.strip())
                                    else:
                                        result[cls] = cell.strip()
                                    break
                        if cls in result and cls != "สติ๊กเกอร์_train":
                            break

    return result

# =========================
# MAIN
# =========================
with open(MAP_FILE, "r", encoding="utf-8") as f:
    records = json.load(f)

for item in records:
    key = item.get("key")
    excel_path = item.get("file_path")

    print(f"\n🔑 KEY: {key}")

    if not excel_path or not os.path.exists(excel_path):
        print("⚠️ ไม่พบไฟล์ Excel")
        continue

    if excel_path.lower().endswith(".xls"):
        excel_path = convert_xls_to_xlsx(excel_path)

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print("❌ เปิด Excel ไม่ได้:", e)
        continue

    ws = next((wb[s] for s in TARGET_SHEETS if s in wb.sheetnames), None)
    if not ws:
        print("⚠️ ไม่พบชีท MASTER SPEC")
        continue

    codes = find_codes(ws)

    # รูปสินค้าสำเร็จ
    codes["รูปสินค้าสำเร็จ_trin"] = key

    # =========================
    # META
    # =========================
    if len(codes) == 1:
        meta = {
            "status": "NO_PACKAGING_CODE",
            "fallback_id": key
        }
    else:
        meta = {"status": "OK"}

    out = {
        "key": key,
        "codes": codes,
        "_meta": meta
    }

    out_path = os.path.join(OUT_DIR, f"{key}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print("✅ saved:", out_path)

print("\n🎉 STEP 3B COMPLETE")
