from openpyxl import load_workbook
import json
import os
import time


def get_latest_modified_folder(parent_dir):
    folders = [os.path.join(parent_dir, d) for d in os.listdir(parent_dir) if os.path.isdir(os.path.join(parent_dir, d))]
    if not folders:
        return None
    latest_folder = max(folders, key=os.path.getmtime)
    return latest_folder

print("="*50)
print(" STEP 1 – LOAD KEY FROM LATEST EXCEL")
print("="*50)

# แสดงโฟลเดอร์ที่แก้ไขล่าสุดใน final_images
latest_folder = get_latest_modified_folder("final_images")
if latest_folder:
    print("📁 โฟลเดอร์ที่แก้ไขล่าสุดใน final_images:", latest_folder)
    print("🕒 เวลาแก้ไขล่าสุด:", time.ctime(os.path.getmtime(latest_folder)))
else:
    print("❌ ไม่พบโฟลเดอร์ใน final_images")

# =====================
# CONFIG
# =====================
# โฟลเดอร์รากที่ใช้ค้นหาไฟล์ Excel ทั้งหมด (ค้นหาแบบ recursive ใต้โฟลเดอร์นี้)
EXCEL_DIR = r"C:\xampp\htdocs\train-ai"      # 👉 ปรับเป็นรากบนไดรฟ์ M: เช่น r"M:\" หรือ r"M:\SUREE"
OUT_DIR = "sheet_data"

KEY_COL = "F"               # คอลัมน์ที่เก็บ key
START_ROW = 3356            # แถวเริ่มต้น
END_ROW = 3376              # แถวสุดท้าย (None = ถึงท้ายชีท)

os.makedirs(OUT_DIR, exist_ok=True)

# =====================
# FIND LATEST EXCEL FILE (RECURSIVE)
# =====================
def get_latest_excel(root_folder):
    excel_files = []

    # เดินทุกโฟลเดอร์ย่อยใต้ root_folder แล้วเก็บไฟล์ Excel ทั้งหมด
    for dirpath, dirnames, filenames in os.walk(root_folder):
        for name in filenames:
            if name.lower().endswith((".xlsx", ".xls")) and not name.startswith("~$"):
                excel_files.append(os.path.join(dirpath, name))

    if not excel_files:
        raise RuntimeError("❌ ไม่พบไฟล์ Excel ในโฟลเดอร์ที่กำหนด (รวมโฟลเดอร์ย่อย)")

    # คืนไฟล์ที่มีเวลาแก้ไขล่าสุด
    return max(excel_files, key=os.path.getmtime)

MASTER_FILE = get_latest_excel(EXCEL_DIR)

print("📄 ใช้ไฟล์ Excel:", MASTER_FILE)
print("🕒 แก้ไขล่าสุด:", time.ctime(os.path.getmtime(MASTER_FILE)))

# =====================
# LOAD EXCEL
# =====================
wb = load_workbook(MASTER_FILE, data_only=True)
ws = wb.active

records = []
last_row = END_ROW if END_ROW else ws.max_row

for row in range(START_ROW, last_row + 1):
    key = ws[f"{KEY_COL}{row}"].value
    if not key:
        continue

    records.append({
        "row": row,
        "key": str(key).strip()
    })

# =====================
# SAVE JSON
# =====================
out_path = os.path.join(OUT_DIR, "keys.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

print("✅ STEP 1 DONE")
print(f"📄 ดึง key ทั้งหมด: {len(records)} รายการ")
print(f"📁 บันทึกที่: {out_path}")
print("="*50)
